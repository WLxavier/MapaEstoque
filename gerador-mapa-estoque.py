import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import time
import datetime
import dates
import cx_Oracle

MESONE, MESTWO, MESTREE, MESFOUR = dates.obter_meses()

# Referencia ao path InstantClient
cx_Oracle.init_oracle_client(lib_dir="./instantclient_21_9")


# Detalhes de conexão
dsn_tns = cx_Oracle.makedsn('***.***.***.***', 1521, service_name='WINT')
connection = cx_Oracle.connect(user='****', password='****', dsn=dsn_tns)

# Função para gerar o mapa
def gerarmapa(linha):
    try:
        CODMARCA = linha['CODMARCA']
        MARCA = linha['MARCA']
        DIRETORIO = linha['DIRETORIO']

        # Obtém a data atual no formato YYYYMMDD
        yyyymmdd = datetime.datetime.today().strftime('%Y%m%d')

        # Alterar a DATA para o dia atual.
        DIRETORIO = DIRETORIO.replace("DATA", yyyymmdd)

        # Ler a consulta SQL do arquivo
        with open('consulta.sql', 'r', encoding='utf-8') as file:
            sql_query = file.read()

        # Criar um cursor
        cursor = connection.cursor()

        # Substituir o valor de :CODMARCA na consulta SQL
        sql_query = sql_query.replace(':CODMARCA', str(CODMARCA))

        # Executar a consulta e obter os resultados
        print(f"Executando consulta para {MARCA}...")
        cursor.execute(sql_query)

        # Obter os resultados e converter em um DataFrame do pandas
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(cursor.fetchall(), columns=columns)
        print("Convertendo resultado em .xlsx")

        # Verifica se o diretório existe, se não, cria
        if not os.path.exists(DIRETORIO):
            os.makedirs(DIRETORIO)

        # Gerar o nome do arquivo Excel
        nome_arquivo = f'{DIRETORIO}\\{MARCA}.xlsx'

        # Abrir o arquivo Excel gerado
        wb = Workbook()
        ws = wb.active


        # Inserir 3 linhas no início do arquivo
        ws.insert_rows(0, 3)

        # Inserir títulos nas células
        ws['A1'] = 'CADASTRO'
        ws['G1'] = 'ESTOQUE'
        ws['K1'] = 'VENDAS'
        ws['A3'] = 'EAN'
        ws['B3'] = 'CÓDIGO'
        ws['C3'] = 'DESCRIÇÃO'
        ws['D3'] = 'MARCA'
        ws['E3'] = 'COD FORNEC.'
        ws['F3'] = 'CX PAD.'
        ws['G3'] = 'FÍSICO'
        ws['H3'] = 'RESERVADO'
        ws['I3'] = 'DISPONÍVEL'
        ws['J3'] = 'PENDÊNCIA'
        ws['K3'] = MESONE
        ws['L3'] = MESTWO
        ws['M3'] = MESTREE
        ws['N3'] = MESFOUR
        ws['O3'] = 'MÉDIA'
        ws['P3'] = 'DIAS DE ESTOQUE'

        # Adicionar fórmula para a soma de campos
        form_G2 = df['FÍSICO'].sum()
        form_H2 = df['RESERVADO'].sum()
        form_I2 = df['DISPONÍVEL'].sum()
        form_J2 = df['PENDÊNCIA'].sum()
        form_K2 = df['MES01'].sum()
        form_L2 = df['MES02'].sum()
        form_M2 = df['MES03'].sum()
        form_N2 = df['MES04'].sum()
        form_O2 = df['MÉDIA'].sum()

        # Inserção de resultado das somas dos DF
        ws['G2'] = form_G2
        ws['H2'] = form_H2
        ws['I2'] = form_I2
        ws['J2'] = form_J2
        ws['K2'] = form_K2
        ws['L2'] = form_L2
        ws['M2'] = form_M2
        ws['N2'] = form_N2
        ws['O2'] = form_O2

        # Aplicar estilo às células
        font_negrito = Font(bold=True, size=12)
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Criando os arquivos de fontes
        fundo_azul_escuro = PatternFill(start_color="002060", end_color="00205B", fill_type="solid")
        fundo_azul_claro = PatternFill(start_color="0070c0", end_color="5B9BD5", fill_type="solid")
        fundo_branco_025 = PatternFill(start_color="bfbfbf", end_color="F2F2F2", fill_type="solid")
        fundo_branco_015 = PatternFill(start_color="d9d9d9", end_color="E6E6E6", fill_type="solid")
        fonte_branca = Font(color="FFFFFF", size=12, bold=True)
        fonte_preta = Font(color="000000", size=12, bold=True)
        fonte_texto = Font(color="000000", size=11, bold=False)


        # Aplicar estilo às células de A1 a A1
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
            for cell in row:
                cell.font = fonte_branca
                cell.fill = fundo_azul_escuro
                
        # Aplicar estilo às células de G1 a J2
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=7, max_col=10):
            for cell in row:
                cell.font = fonte_branca
                cell.fill = fundo_azul_claro
                
        # Aplicar estilo às células de K1 a P2
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=11, max_col=16):
            for cell in row:
                cell.font = fonte_branca
                cell.fill = fundo_azul_escuro
                
        # Aplicar estilo às células de A3 a P3
        for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=16):
            for cell in row:
                cell.font = fonte_preta
                cell.fill = fundo_branco_025

        # Adicionar os valores do DataFrame ao Excel
        for r_idx, row in df.iterrows():
            ws.append(row.tolist())

        # Aplicar estilo às células de I4 pra baixo
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=9, max_col=9):
            for cell in row:
                cell.fill = fundo_branco_015
                cell.font = fonte_texto

        # Transforma em número as células de A4 pra baixo
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '0'

        # Centralizar todas as células
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.alignment = alignment

        # Loop para ajustar automaticamente o tamanho das colunas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.01
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Defina a largura da coluna A manualmente
        ws.column_dimensions['A'].width = 15

        # Aplicar estilo às células de A1 a P3
        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=16):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Mesclar as células
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=10)
        ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=16)

        # Adicionar borda em todas as células ocupadas
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )


        # Ativar o filtro para as células de A3 a P3
        ws.auto_filter.ref = 'A3:P3'

        # Fechar o cursor
        cursor.close()

        # Salvar o arquivo
        wb.save(nome_arquivo)

        print(f"Arquivo {MARCA} gerado com sucesso!")

    except Exception as e:
        print(f"Erro ao gerar o mapa: {e}")

# Função para carregar dados e agendar envios (modificada)
def carregar_dados_e_agendar():
    try:
        # Inicializar a última modificação
        carregar_dados_e_agendar.ultima_modificacao = os.path.getmtime('./gerador-mapa-de-estoque.xlsx')

        # Ler os dados da planilha Excel
        dados = pd.read_excel('./gerador-mapa-de-estoque.xlsx')

        for index, linha in dados.iterrows():
            horario = linha['HORARIO'].strftime('%H:%M')

            # Verifica se é hora de executar para esta linha
            if horario == time.strftime('%H:%M'):
                try:
                    # Executar a função para gerar o mapa
                    gerarmapa(linha)
                except Exception as e:
                    print(f"Erro ao gerar o arquivo: {e}")
    except Exception as e:
        print(f"Erro ao carregar dados da planilha: {e}")

# Loop infinito para verificar periodicamente
while True:
    try:
        carregar_dados_e_agendar()
        time.sleep(60)  # Aguarda 60 segundos antes de verificar novamente
    except Exception as e:
        print(f"Erro durante a execução do loop principal: {e}")
